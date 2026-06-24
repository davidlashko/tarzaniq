"""Excel export (and re-import) for one day.

The workbook is the portable archive: Photos/Subjects/Engagements are
the human-readable dataset, and the Meta sheet carries the complete
day record as chunked JSON so a lost database can be rebuilt from the
exports alone (Settings -> Import in the dashboard).
"""

import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from . import APP_VERSION
from .stats import fmt_dur, fmt_pct

GREEN = "1F6F43"
DARK = "12331F"
BANANA = "FFD23F"
SAND = "F4E9D0"
BROWN = "6B4A2F"

F_BASE = Font(name="Arial", size=10)
F_HDR = Font(name="Arial", size=10, bold=True, color="FFFFFF")
F_SECTION = Font(name="Arial", size=11, bold=True, color="FFFFFF")
F_TITLE = Font(name="Arial", size=14, bold=True, color=DARK)

FILL_HDR = PatternFill("solid", start_color=GREEN)
FILL_SECTION = PatternFill("solid", start_color=BROWN)
FILL_TITLE = PatternFill("solid", start_color=BANANA)
FILL_STRIPE = PatternFill("solid", start_color=SAND)

MAX_CELL = 30000  # stay safely under Excel's 32,767-char cell limit


def _autosize(ws, cap=58):
    widths = {}
    for row in ws.iter_rows():
        for c in row:
            if c.value is None:
                continue
            L = len(str(c.value))
            widths[c.column] = min(max(widths.get(c.column, 0), L + 2), cap)
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _header(ws, row, labels):
    for i, lab in enumerate(labels, start=1):
        c = ws.cell(row=row, column=i, value=lab)
        c.font = F_HDR
        c.fill = FILL_HDR
        c.alignment = Alignment(horizontal="center")
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def export_day(day_record, out_path: Path):
    """day_record: the dict handed to db.commit_day (plus 'stats')."""
    st = day_record["stats"]
    wb = Workbook()

    # ---------------------------------------------------------- Summary
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = f"TarzanIQ — {day_record['date']}  {day_record['place']}  {day_record['employee']}"
    ws["A1"].font = F_TITLE
    ws["A1"].fill = FILL_TITLE
    ws.merge_cells("A1:C1")

    r = [3]  # mutable row cursor

    def section(name):
        c = ws.cell(row=r[0], column=1, value=name)
        c.font = F_SECTION
        c.fill = FILL_SECTION
        ws.merge_cells(start_row=r[0], start_column=1,
                       end_row=r[0], end_column=3)
        r[0] += 1

    cells = {}

    def kv(label, value, pretty=None, key=None, fmt=None):
        ws.cell(row=r[0], column=1, value=label).font = F_BASE
        c = ws.cell(row=r[0], column=2, value=value)
        c.font = F_BASE
        if fmt:
            c.number_format = fmt
        if pretty is not None:
            ws.cell(row=r[0], column=3, value=pretty).font = F_BASE
        if r[0] % 2 == 0:
            for col in (1, 2, 3):
                ws.cell(row=r[0], column=col).fill = FILL_STRIPE
        if key:
            cells[key] = f"B{r[0]}"
        r[0] += 1

    section("DAY")
    kv("Date", day_record["date"])
    kv("Weekday", day_record["weekday"])
    kv("Place", day_record["place"])
    kv("Photographer", day_record["employee"])
    kv("Source folder", day_record.get("source_folder") or "")
    kv("Photos total", st["photos_total"], key="photos")
    kv("Photos with subject in focus", st["photos_focus"])
    kv("Air shots (no subject)", st["photos_air"])
    kv("Files skipped (non-JPEG)", st["skipped_files"])
    kv("Photos missing EXIF time", st["missing_exif"])
    kv("Suspected deletions", st["suspected_deletions"])
    kv("First shot", st["first_shot"] or "")
    kv("Last shot", st["last_shot"] or "")
    kv("Hours on street", round(st["span_s"] / 3600, 2),
       fmt_dur(st["span_s"]), key="span")
    kv("Shooting hours", round(st["shoot_s"] / 3600, 2),
       fmt_dur(st["shoot_s"]), key="shoot_h")
    kv("Break time", round(st["break_s"] / 3600, 2), fmt_dur(st["break_s"]))
    kv("Breaks taken", st["breaks_n"])

    section("HUNT")
    kv("Cold shoots (approach events)", st["cold_events"], key="cold_e")
    kv("Cold persons (unique subjects)", st["cold_persons"], key="cold_p")
    kv("Cold persons / shooting hr",
       f"=IF({cells['shoot_h']}=0,\"-\",{cells['cold_p']}/{cells['shoot_h']})",
       fmt="0.00")
    kv("Avg group size",
       round(st["avg_group_size"], 2) if st["avg_group_size"] else "-")
    kv("% approaches that were groups", st["pct_group_approaches"],
       fmt_pct(st["pct_group_approaches"]), fmt="0.0%")
    kv("Avg hunting time (between marks)", 
       round(st["hunting_avg_s"], 1) if st["hunting_avg_s"] else "-",
       fmt_dur(st["hunting_avg_s"]))
    kv("Longest dry spell", round(st["dry_spell_s"], 1),
       fmt_dur(st["dry_spell_s"]))
    kv("Re-approached subjects", st["reapproaches"])

    section("CLOSE")
    kv("Warm persons (converted)", st["warm_persons"], key="warm_p")
    kv("CONVERSION (warm / cold persons)",
       f"=IF({cells['cold_p']}=0,\"-\",{cells['warm_p']}/{cells['cold_p']})",
       fmt="0.0%")
    kv("Warm persons / shooting hr",
       f"=IF({cells['shoot_h']}=0,\"-\",{cells['warm_p']}/{cells['shoot_h']})",
       fmt="0.00")
    kv("Avg pitch time (cold end -> first warm)",
       round(st["pitch_avg_s"], 1) if st["pitch_avg_s"] else "-",
       fmt_dur(st["pitch_avg_s"]))
    kv("Solo conversion", st["solo_conv"], fmt_pct(st["solo_conv"]), fmt="0.0%")
    kv("Group conversion", st["group_conv"], fmt_pct(st["group_conv"]), fmt="0.0%")
    kv("Hot streak (consecutive converting approaches)", st["hot_streak"])

    section("HOLD")
    kv("Warm sessions", st["warm_sessions_n"])
    kv("Avg warm shoot duration",
       round(st["warm_dur_avg_s"], 1) if st["warm_dur_avg_s"] else "-",
       fmt_dur(st["warm_dur_avg_s"]))
    kv("Avg photos per warm shoot",
       round(st["warm_photos_avg"], 1) if st["warm_photos_avg"] else "-")
    kv("Avg poses per warm shoot (est.)",
       round(st["poses_avg"], 1) if st["poses_avg"] else "-")
    kv("Total time in warm shoots", round(st["warm_time_total_s"], 1),
       fmt_dur(st["warm_time_total_s"]))

    section("RATES")
    kv("Photos / shooting hr",
       f"=IF({cells['shoot_h']}=0,\"-\",{cells['photos']}/{cells['shoot_h']})",
       fmt="0.0")
    kv("Avg faces in frame (focus photos)",
       round(st["avg_faces_in_frame"], 2) if st["avg_faces_in_frame"] else "-")

    section("PEOPLE")
    for g, n in sorted(st["gender_count"].items()):
        tot = max(st["cold_persons"], 1)
        kv(f"Gender {g}", n, f"{n / tot * 100:.0f}% of subjects")
    for g, n in sorted(st["gender_warm"].items()):
        base = st["gender_count"].get(g, 0)
        kv(f"Conversion — gender {g}",
           (n / base) if base else "-",
           f"{n} of {base}", fmt="0.0%")
    for a, n in sorted(st["age_count"].items()):
        tot = max(st["cold_persons"], 1)
        kv(f"Age {a}", n, f"{n / tot * 100:.0f}% of subjects")
    for a, n in sorted(st["age_warm"].items()):
        base = st["age_count"].get(a, 0)
        kv(f"Conversion — age {a}",
           (n / base) if base else "-",
           f"{n} of {base}", fmt="0.0%")

    section("MONEY (self-reported, daily lump)")
    kv("Cash", day_record.get("money_cash"), key="cash")
    kv("Card", day_record.get("money_card"), key="card")
    kv("Total", f"=IF(AND({cells['cash']}=\"\",{cells['card']}=\"\"),\"-\","
                f"N({cells['cash']})+N({cells['card']}))")

    _autosize(ws)
    ws.column_dimensions["A"].width = 42

    # ---------------------------------------------------------- Photos
    ws = wb.create_sheet("Photos")
    _header(ws, 1, ["filename", "seq", "time", "kind", "focus_faces",
                    "rejected_faces", "subjects", "flags", "detail_json"])
    for i, p in enumerate(day_record["photos"], start=2):
        ws.cell(row=i, column=1, value=p["filename"]).font = F_BASE
        ws.cell(row=i, column=2, value=p.get("seq")).font = F_BASE
        ws.cell(row=i, column=3, value=p["t"]).font = F_BASE
        ws.cell(row=i, column=4, value=p.get("kind")).font = F_BASE
        ws.cell(row=i, column=5, value=p.get("n_focus", 0)).font = F_BASE
        ws.cell(row=i, column=6, value=p.get("n_rejected", 0)).font = F_BASE
        ws.cell(row=i, column=7,
                value=",".join(f"S{s + 1}" for s in p.get("subjects", []))).font = F_BASE
        ws.cell(row=i, column=8, value=";".join(p.get("flags", []))).font = F_BASE
        ws.cell(row=i, column=9,
                value=json.dumps(p.get("detail", {}))[:MAX_CELL]).font = F_BASE
    _autosize(ws, cap=40)

    # ---------------------------------------------------------- Subjects
    ws = wb.create_sheet("Subjects")
    _header(ws, 1, ["subject", "gender", "gender_conf", "age_bucket",
                    "age_est", "photos", "did_warm", "pitch_s",
                    "warm_sessions", "warm_photos", "warm_duration_s",
                    "poses_est", "reapproached", "first_seen", "last_seen"])
    for i, s in enumerate(day_record["subjects"], start=2):
        vals = [f"S{s['local_id'] + 1}", s.get("gender"), s.get("gender_conf"),
                s.get("age_bucket"), s.get("age_est"), s.get("photo_count"),
                "yes" if s.get("did_warm") else "no",
                round(s["pitch_s"], 1) if s.get("pitch_s") is not None else None,
                s.get("warm_sessions"), s.get("warm_photos"),
                round(s.get("warm_duration_s") or 0, 1), s.get("poses_est"),
                "yes" if s.get("reapproached") else "no",
                s.get("first_seen"), s.get("last_seen")]
        for j, v in enumerate(vals, start=1):
            ws.cell(row=i, column=j, value=v).font = F_BASE
    _autosize(ws)

    # ---------------------------------------------------------- Engagements
    ws = wb.create_sheet("Engagements")
    _header(ws, 1, ["kind", "start", "end", "duration_s", "members",
                    "n_members", "n_converted", "photos", "poses",
                    "reapproach"])
    for i, e in enumerate(day_record["engagements"], start=2):
        mem = e.get("members")
        if isinstance(mem, list):
            mem = ",".join(f"S{m + 1}" for m in mem)
        elif isinstance(mem, int):
            mem = f"S{mem + 1}"
        vals = [e["kind"], e.get("start"), e.get("end"),
                round(e.get("duration_s") or 0, 1), mem, e.get("n_members"),
                e.get("n_converted"), e.get("photos"), e.get("poses"),
                "yes" if e.get("reapproach") else ""]
        for j, v in enumerate(vals, start=1):
            ws.cell(row=i, column=j, value=v).font = F_BASE
    _autosize(ws)

    # ---------------------------------------------------------- Hourly
    ws = wb.create_sheet("Hourly")
    _header(ws, 1, ["hour", "shooting_min", "cold_persons", "warm_persons",
                    "conversion"])
    for i, hrow in enumerate(st.get("hourly", []), start=2):
        ws.cell(row=i, column=1, value=f"{hrow['hour']:02d}:00").font = F_BASE
        ws.cell(row=i, column=2,
                value=round(hrow["shoot_s"] / 60, 1)).font = F_BASE
        ws.cell(row=i, column=3, value=hrow["cold_p"]).font = F_BASE
        ws.cell(row=i, column=4, value=hrow["warm_p"]).font = F_BASE
        c = ws.cell(row=i, column=5,
                    value=f"=IF(C{i}=0,\"-\",D{i}/C{i})")
        c.font = F_BASE
        c.number_format = "0.0%"
    _autosize(ws)

    # ---------------------------------------------------------- Meta
    ws = wb.create_sheet("Meta")
    ws["A1"] = "TarzanIQ machine data — do not edit (used for re-import)"
    ws["A1"].font = Font(name="Arial", size=10, bold=True, color=BROWN)
    ws["A2"] = "app_version"
    ws["B2"] = APP_VERSION
    ws["A3"] = "exported_at"
    ws["B3"] = datetime.now().isoformat()
    ws["A4"] = "params"
    ws["B4"] = json.dumps(day_record["params"])
    blob = json.dumps(day_record, default=str)
    chunks = [blob[i:i + MAX_CELL] for i in range(0, len(blob), MAX_CELL)]
    ws["A5"] = "chunks"
    ws["B5"] = len(chunks)
    for i, ch in enumerate(chunks):
        ws.cell(row=6 + i, column=1, value=f"data_{i}")
        ws.cell(row=6 + i, column=2, value=ch)
    ws.column_dimensions["A"].width = 16
    for row in ws.iter_rows():
        for c in row:
            if c.font is None or c.font.name != "Arial":
                c.font = F_BASE

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def import_day(xlsx_path: Path):
    """Rebuild a day_record from a TarzanIQ export's Meta sheet."""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    if "Meta" not in wb.sheetnames:
        raise ValueError("Not a TarzanIQ export (no Meta sheet)")
    ws = wb["Meta"]
    rows = {str(r[0].value): r[1].value for r in
            ws.iter_rows(min_row=2, max_col=2) if r[0].value}
    n = int(rows.get("chunks", 0))
    blob = "".join(str(rows.get(f"data_{i}", "")) for i in range(n))
    rec = json.loads(blob)
    wb.close()
    return rec
